"""Tests for jobs/excel_replenishment_job.py — replenish a client's planilla.

Mirrors the odoo_replenishment shape: prepare reads the system of record (here
the client's Excel file), run plans the restock and STAGES the write-back as a
dry-run changeset through the safe-staging plane, and the outcome is >=2 ranked
executable options. Nothing is ever written without an approval + apply.
"""

from __future__ import annotations

import pytest
from openpyxl import Workbook, load_workbook

from jobs import excel_replenishment_job as job
from src import writeback
from src.guided import OPTIONS, passed_guided

SHEET = "Stock Bodega"


def _make_planilla(path, *, demand_column=False, order_column=False):
    """Client-style planilla: title row, Spanish headers at row 3, own formulas."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws["A1"] = "INVENTARIO BODEGA CENTRAL"
    headers = ["Codigo", "Descripcion", "Stock", "Punto Reorden"]
    if demand_column:
        headers.append("Demanda Semanal")
    if order_column:
        headers.append("Pedir (Linchpin)")
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    rows = [
        ("SKU-001", "Tornillo 3mm", 42, 50, 10.0),
        ("SKU-002", "Tuerca 3mm", 130, 80, 12.0),
        ("SKU-003", "Arandela", 8, 25, 4.0),
    ]
    for r, (code, desc, stock, rop, demand) in enumerate(rows, 4):
        ws.cell(row=r, column=1, value=code)
        ws.cell(row=r, column=2, value=desc)
        ws.cell(row=r, column=3, value=stock)
        ws.cell(row=r, column=4, value=rop)
        if demand_column:
            ws.cell(row=r, column=5, value=demand)
    wb.save(path)
    return path


@pytest.fixture
def planilla(tmp_path):
    return _make_planilla(tmp_path / "planilla.xlsx")


# ---- prepare: sheet + column auto-detection ---------------------------------------

def test_prepare_autodetects_sheet_and_spanish_columns(planilla):
    payload = job.prepare(str(planilla), {})
    assert payload["sheet"] == SHEET
    assert payload["mode"] == "reorder-point"
    skus = [row.sku for row in payload["rows"]]
    assert skus == ["SKU-001", "SKU-002", "SKU-003"]
    assert payload["rows"][0].on_hand == 42
    assert payload["rows"][0].reorder_point == 50


def test_prepare_prefers_demand_mode_when_demand_column_present(tmp_path):
    p = _make_planilla(tmp_path / "d.xlsx", demand_column=True)
    payload = job.prepare(str(p), {})
    assert payload["mode"] == "demand-cover"
    assert payload["rows"][0].demand_per_period == 10.0


def test_prepare_respects_explicit_column_params(planilla):
    payload = job.prepare(str(planilla), {"sku_column": "Codigo", "stock_column": "Stock",
                                          "rop_column": "Punto Reorden", "sheet": SHEET})
    assert payload["sheet"] == SHEET
    assert len(payload["rows"]) == 3


def test_prepare_fails_clearly_without_sku_column(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Cosa", "Cantidad"])
    ws.append(["x", 1])
    f = tmp_path / "bad.xlsx"
    wb.save(f)
    with pytest.raises(ValueError, match="SKU"):
        job.prepare(str(f), {})


def test_prepare_fails_clearly_without_rop_or_demand(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Codigo", "Stock"])
    ws.append(["SKU-1", 5])
    f = tmp_path / "no_signal.xlsx"
    wb.save(f)
    with pytest.raises(ValueError, match="reorder point|demand"):
        job.prepare(str(f), {})


# ---- run: plan + staged changeset ---------------------------------------------------

def test_run_reorder_point_mode_orders_up_to_factor(planilla):
    payload = job.prepare(str(planilla), {})
    report = job.run(payload)
    # SKU-001: 42 < 50 -> order up to 2*50 => +58; SKU-002: 130 >= 80 -> 0;
    # SKU-003: 8 < 25 -> +42.
    assert report.restock == {"SKU-001": 58.0, "SKU-003": 42.0}
    assert report.n_restock == 2 and report.n_skus == 3
    assert report.total_restock == 100.0


def test_run_demand_cover_mode_targets_cover_periods(tmp_path):
    p = _make_planilla(tmp_path / "d.xlsx", demand_column=True)
    payload = job.prepare(str(p), {})
    report = job.run(payload, cover_periods=8.0)
    # SKU-001: target 80 vs 42 -> +38; SKU-002: 96 vs 130 -> 0; SKU-003: 32 vs 8 -> +24.
    assert report.restock == {"SKU-001": 38.0, "SKU-003": 24.0}
    assert report.mode == "demand-cover"


def test_run_stages_changeset_with_new_order_column(planilla):
    payload = job.prepare(str(planilla), {})
    report = job.run(payload)
    cs = report.changeset
    assert cs is not None and cs.risk_tier == writeback.TIER_REVERSIBLE
    edits = {c.field: c.after for c in cs.changes}
    # New column E: header at the header row + one qty per restocked SKU.
    assert edits["E3"] == "Pedir (Linchpin)"
    assert edits["E4"] == 58.0 and edits["E6"] == 42.0
    assert all(c.before is None for c in cs.changes)  # dry-run: nothing written yet
    assert load_workbook(planilla)[SHEET]["E4"].value is None  # file untouched


def test_run_reuses_existing_order_column(tmp_path):
    p = _make_planilla(tmp_path / "o.xlsx", order_column=True)
    payload = job.prepare(str(p), {})
    report = job.run(payload)
    edits = {c.field: c.after for c in report.changeset.changes}
    assert "E3" not in edits  # header already exists -> not re-written
    assert edits["E4"] == 58.0


def test_run_no_restock_needed_stages_nothing(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.append(["Codigo", "Stock", "Punto Reorden"])
    ws.append(["SKU-1", 100, 10])
    f = tmp_path / "full.xlsx"
    wb.save(f)
    report = job.run(job.prepare(str(f), {}))
    assert report.restock == {} and report.changeset is None
    assert report.outcome.status == OPTIONS  # still a protected, ranked outcome


# ---- outcome contract ----------------------------------------------------------------

def test_outcome_offers_ranked_executable_options(planilla):
    report = job.run(job.prepare(str(planilla), {}))
    out = report.outcome
    assert out.status == OPTIONS
    assert len(out.options) >= 2
    assert sum(1 for o in out.options if o.recommended) == 1
    assert all(o.action for o in out.options)
    assert passed_guided(out)


# ---- the loop actually closes: approve + apply the staged changeset -------------------

def test_staged_changeset_applies_to_the_real_file_with_approval(planilla):
    payload = job.prepare(str(planilla), {})
    report = job.run(payload)
    store = payload["store"]
    approval = writeback.approve(report.changeset, "operator")
    result = writeback.apply(store, report.changeset, approval=approval)
    assert result.applied
    ws = load_workbook(planilla)[SHEET]
    assert ws["E3"].value == "Pedir (Linchpin)"
    assert ws["E4"].value == 58.0
    assert ws["A1"].value == "INVENTARIO BODEGA CENTRAL"  # client content intact
    # And it is rollback-able, honoring the writeback contract end to end.
    store.rollback(report.changeset.idempotency_key)
    assert load_workbook(planilla)[SHEET]["E4"].value is None


# ---- verify / deliverables -------------------------------------------------------------

def test_verify_passes_on_good_report(planilla):
    report = job.run(job.prepare(str(planilla), {}))
    assert job.verify(report) == []


def test_verify_flags_empty_planilla(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.append(["Codigo", "Stock", "Punto Reorden"])
    f = tmp_path / "empty.xlsx"
    wb.save(f)
    with pytest.raises(ValueError, match="no SKU rows"):
        job.prepare(str(f), {})


def test_write_operational_emits_csv(planilla, tmp_path):
    report = job.run(job.prepare(str(planilla), {}))
    written = job.write_operational(report, tmp_path / "out", "Acme")
    assert written["csv"].exists()
    text = written["csv"].read_text(encoding="utf-8")
    assert "SKU-001" in text and "58" in text


def test_build_deck_writes_deliverable(planilla, tmp_path):
    report = job.run(job.prepare(str(planilla), {}))
    deck = job.build_deck(report, client="Acme", citations=("Vandeput (2020), ch. 2",))
    files = deck.write_all(tmp_path / "deck")
    assert any(p.exists() for p in files.values())
