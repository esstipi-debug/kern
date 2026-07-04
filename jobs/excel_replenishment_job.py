"""Excel replenishment agent job: read a client's planilla -> plan -> staged write-back.

The Excel twin of ``jobs/odoo_job.py``: the client's spreadsheet IS the system of
record. ``prepare`` reads the planilla (auto-detecting the sheet, the header row and
the SKU / stock / reorder-point / demand columns, Spanish or English, accents
folded), ``run`` plans the restock and STAGES the recommended order quantities as a
dry-run changeset through the safe-staging plane (``src/connectors/excel.py`` —
drift check, atomic write, backup, rollback), and the result is presented as >=2
ranked executable options honouring the never-unprotected contract. Nothing touches
the client's file until a human approves and applies the staged changeset.

Two planning modes, picked from what the planilla actually has:
- ``demand-cover``: a demand column exists -> order up to ``cover_periods`` of demand
  (same target logic as ``src/connectors/replenish.py``);
- ``reorder-point``: only a reorder-point column exists -> classic min/max: order when
  on-hand < ROP, up to ``order_up_to_factor * ROP``.
"""

from __future__ import annotations

import unicodedata
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src import writeback
from src.connectors.excel import ExcelWorkbookStore
from src.deliverable import DataSource, Deliverable, Finding, Kpi
from src.export import write_summary_csv
from src.guided import ExecutionOption, GuidedOutcome, as_options, verify_guided

_EXCEL_SUFFIXES = (".xlsx", ".xlsm", ".xltx", ".xltm")
_HEADER_SCAN_ROWS = 20
DEFAULT_ORDER_COLUMN = "Pedir (Linchpin)"

_SKU_CANDIDATES = {
    "sku", "skus", "codigo", "codigo sku", "code", "item", "producto", "product",
    "product id", "product_id", "referencia", "ref", "articulo", "material",
}
_STOCK_CANDIDATES = {
    "stock", "on hand", "on-hand", "on_hand", "existencia", "existencias", "cantidad",
    "qty", "quantity", "inventario", "disponible", "saldo", "stock actual",
}
_ROP_CANDIDATES = {
    "punto reorden", "punto de reorden", "reorder point", "rop", "minimo", "min",
    "stock minimo", "reorden", "punto pedido", "punto de pedido",
}
_DEMAND_CANDIDATES = {
    "demanda", "demanda semanal", "demanda mensual", "demand", "weekly demand",
    "venta promedio", "ventas promedio", "avg sales", "average sales", "forecast",
    "pronostico", "consumo", "consumo promedio",
}


def _fold(label: object) -> str:
    """Lowercased, accent-stripped, whitespace-collapsed header label."""
    text = unicodedata.normalize("NFKD", str(label)).encode("ascii", "ignore").decode("ascii")
    return " ".join(text.lower().split())


def _num(value: object) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return None


@dataclass(frozen=True)
class PlanillaRow:
    row: int
    sku: str
    on_hand: float
    reorder_point: float | None
    demand_per_period: float | None


@dataclass(frozen=True)
class ReplenishmentLine:
    sku: str
    on_hand: float
    target: float
    restock_qty: float


@dataclass(frozen=True)
class ExcelReplenishmentReport:
    filename: str
    sheet: str
    mode: str  # "demand-cover" | "reorder-point"
    lines: tuple[ReplenishmentLine, ...]
    restock: dict[str, float]
    n_skus: int
    n_restock: int
    total_restock: float
    cover_periods: float
    order_up_to_factor: float
    changeset: writeback.Changeset | None
    outcome: GuidedOutcome
    summary: str


def _find_columns(ws, wanted: dict[str, set[str]]) -> tuple[int, dict[str, str]] | None:
    """First row (within the scan window) containing a SKU-candidate header.

    Returns (header_row, {role: column_letter}) or None. ``wanted`` maps role ->
    folded candidate labels; the SKU role is what anchors the header row.
    """
    for row in ws.iter_rows(min_row=1, max_row=_HEADER_SCAN_ROWS):
        labels = {_fold(c.value): c.column_letter for c in row if c.value is not None}
        sku_hits = [labels[c] for c in wanted["sku"] if c in labels]
        if not sku_hits:
            continue
        found = {"sku": sku_hits[0]}
        for role in ("stock", "rop", "demand", "order"):
            hits = [labels[c] for c in wanted[role] if c in labels]
            if hits:
                found[role] = hits[0]
        return row[0].row, found
    return None


def prepare(data_path: str | None, params: dict | None = None) -> dict:
    """Read the client's planilla and resolve where everything lives.

    ``params`` overrides beat auto-detection: ``sheet``, ``sku_column``,
    ``stock_column``, ``rop_column``, ``demand_column``, ``order_column``.
    """
    params = params or {}
    if not data_path:
        raise ValueError("an Excel file (.xlsx/.xlsm) is required")
    path = Path(data_path)
    if path.suffix.lower() not in _EXCEL_SUFFIXES:
        raise ValueError(
            f"excel_replenishment needs an Excel file (.xlsx/.xlsm), got {path.suffix!r} - "
            "for CSV demand data use the inventory_optimization tool instead"
        )
    wb = load_workbook(path)  # raises FileNotFoundError on a missing file

    order_column = str(params.get("order_column", DEFAULT_ORDER_COLUMN))
    wanted = {
        "sku": {_fold(params["sku_column"])} if params.get("sku_column") else _SKU_CANDIDATES,
        "stock": {_fold(params["stock_column"])} if params.get("stock_column") else _STOCK_CANDIDATES,
        "rop": {_fold(params["rop_column"])} if params.get("rop_column") else _ROP_CANDIDATES,
        "demand": {_fold(params["demand_column"])} if params.get("demand_column") else _DEMAND_CANDIDATES,
        "order": {_fold(order_column)},
    }

    sheet_names = [params["sheet"]] if params.get("sheet") else wb.sheetnames
    if params.get("sheet") and params["sheet"] not in wb.sheetnames:
        raise ValueError(f"sheet {params['sheet']!r} not found (sheets: {', '.join(wb.sheetnames)})")

    resolved = None
    for name in sheet_names:
        hit = _find_columns(wb[name], wanted)
        if hit is not None:
            resolved = (name, *hit)
            break
    if resolved is None:
        raise ValueError(
            "could not find a SKU/product column in the first rows of any sheet - "
            "name it (e.g. 'Codigo'/'SKU') or pass sku_column/sheet explicitly"
        )
    sheet, header_row, cols = resolved

    if "stock" not in cols:
        raise ValueError(f"no stock/on-hand column found next to the SKU column in {sheet!r}")
    if "rop" not in cols and "demand" not in cols:
        raise ValueError(
            f"{sheet!r} has neither a reorder point column (e.g. 'Punto Reorden') nor a "
            "demand column (e.g. 'Demanda Semanal') - one of the two is needed to plan"
        )
    mode = "demand-cover" if "demand" in cols else "reorder-point"

    ws = wb[sheet]
    rows: list[PlanillaRow] = []
    for r in range(header_row + 1, ws.max_row + 1):
        sku_raw = ws[f"{cols['sku']}{r}"].value
        on_hand = _num(ws[f"{cols['stock']}{r}"].value)
        if sku_raw is None or on_hand is None:
            continue  # blank/separator rows, or non-numeric stock (e.g. a formula)
        rows.append(PlanillaRow(
            row=r,
            sku=str(sku_raw).strip(),
            on_hand=on_hand,
            reorder_point=_num(ws[f"{cols['rop']}{r}"].value) if "rop" in cols else None,
            demand_per_period=_num(ws[f"{cols['demand']}{r}"].value) if "demand" in cols else None,
        ))
    if not rows:
        raise ValueError(f"no SKU rows found under the {sheet!r} headers")

    order_exists = "order" in cols
    order_letter = cols["order"] if order_exists else get_column_letter(ws.max_column + 1)
    return {
        "store": ExcelWorkbookStore(path),
        "filename": path.name,
        "sheet": sheet,
        "header_row": header_row,
        "mode": mode,
        "rows": tuple(rows),
        "order_col_name": order_column,
        "order_col_letter": order_letter,
        "order_col_exists": order_exists,
    }


def _build_outcome(n_restock: int, total: float, filename: str,
                   changeset: writeback.Changeset | None) -> GuidedOutcome:
    """The plan as >=2 ranked, executable options (first = recommended default)."""
    if n_restock > 0:
        key = changeset.idempotency_key if changeset is not None else "n/a"
        options = [
            ExecutionOption(
                label="Apply the staged order quantities to the planilla",
                summary=f"Write the recommended order column into {filename} for {n_restock} SKU(s) "
                        "after your approval - atomic, backed up, rollback-able.",
                score=3.0, recommended=True,
                action=f"writeback.approve + apply the staged changeset (key={key}, reversible)",
                tradeoffs="lowest touch; the planilla stays the single source of truth",
            ),
            ExecutionOption(
                label="Export the plan for review",
                summary="Hand off the replenishment plan without touching the client's file.",
                score=1.0,
                action="export the replenishment plan (no write-back)",
                tradeoffs="zero risk; manual follow-up",
            ),
        ]
        summary = f"{n_restock} SKU(s) below target ({total:,.0f} units short): choose how to act."
    else:
        options = [
            ExecutionOption(
                label="Hold - every SKU is above target",
                summary="No replenishment is needed now; nothing to write.",
                score=3.0, recommended=True,
                action="monitor; no write-back needed", tradeoffs="no cost",
            ),
            ExecutionOption(
                label="Tighten the target",
                summary="Lower the cover / order-up-to factor to release working capital.",
                score=2.0,
                action="re-run with a lower cover_periods / order_up_to_factor",
                tradeoffs="frees cash; less buffer",
            ),
        ]
        summary = "All SKUs above target: choose how to proceed."
    return as_options(summary, options)


def run(
    payload: dict,
    *,
    cover_periods: float = 8.0,
    order_up_to_factor: float = 2.0,
    idempotency_key: str = "excel-replenish-1",
) -> ExcelReplenishmentReport:
    """Plan the restock and stage the write-back as a dry-run changeset."""
    if cover_periods <= 0 or order_up_to_factor <= 0:
        raise ValueError("cover_periods and order_up_to_factor must be > 0")
    mode: str = payload["mode"]
    lines: list[ReplenishmentLine] = []
    restock: dict[str, tuple[int, float]] = {}  # sku -> (sheet row, qty)
    for row in payload["rows"]:
        if mode == "demand-cover":
            target = (row.demand_per_period or 0.0) * cover_periods
            qty = max(0.0, round(target - row.on_hand, 1))
        else:
            rop = row.reorder_point or 0.0
            target = rop * order_up_to_factor
            qty = max(0.0, round(target - row.on_hand, 1)) if row.on_hand < rop else 0.0
        lines.append(ReplenishmentLine(row.sku, row.on_hand, target, qty))
        if qty > 0:
            restock[row.sku] = (row.row, qty)

    changeset = None
    if restock:
        letter = payload["order_col_letter"]
        cells: dict[str, object] = {}
        if not payload["order_col_exists"]:
            cells[f"{letter}{payload['header_row']}"] = payload["order_col_name"]
        for _sku, (r, qty) in restock.items():
            cells[f"{letter}{r}"] = qty
        changeset = writeback.stage(
            payload["store"], f"excel:{payload['filename']}", {payload["sheet"]: cells},
            risk_tier=writeback.TIER_REVERSIBLE, idempotency_key=idempotency_key,
            reason=f"replenish {len(restock)} SKU(s) to target ({mode})",
        )

    flat_restock = {sku: qty for sku, (_r, qty) in restock.items()}
    total = sum(flat_restock.values())
    summary = (
        f"Read {len(lines)} SKU(s) from {payload['filename']} ({payload['sheet']}, {mode}); "
        f"{len(flat_restock)} below target ({total:,.0f} units short), staged as a dry-run."
        if flat_restock else
        f"Read {len(lines)} SKU(s) from {payload['filename']} ({payload['sheet']}, {mode}); "
        "all above target - nothing to write."
    )
    return ExcelReplenishmentReport(
        filename=payload["filename"],
        sheet=payload["sheet"],
        mode=mode,
        lines=tuple(lines),
        restock=flat_restock,
        n_skus=len(lines),
        n_restock=len(flat_restock),
        total_restock=total,
        cover_periods=cover_periods,
        order_up_to_factor=order_up_to_factor,
        changeset=changeset,
        outcome=_build_outcome(len(flat_restock), total, payload["filename"], changeset),
        summary=summary,
    )


def verify(report: ExcelReplenishmentReport) -> list[str]:
    """QA gate: protected outcome, real rows, and staging consistent with the plan."""
    issues = list(verify_guided(report.outcome))
    if report.n_skus == 0:
        issues.append("no SKU rows read from the planilla")
    if any(ln.restock_qty < 0 for ln in report.lines):
        issues.append("negative restock quantity in the plan")
    if report.restock and report.changeset is None:
        issues.append("restock planned but no changeset was staged")
    if not report.restock and report.changeset is not None:
        issues.append("a changeset was staged with nothing to restock")
    return issues


def write_operational(report: ExcelReplenishmentReport, out_dir: str | Path,
                      client: str = "Client") -> dict[str, Path]:
    """Machine-readable deliverable: one row per SKU with on-hand / target / restock."""
    d = Path(out_dir)
    d.mkdir(parents=True, exist_ok=True)
    rows = [
        {
            "sku": ln.sku,
            "on_hand": round(ln.on_hand, 2),
            "target": round(ln.target, 2),
            "restock_qty": round(ln.restock_qty, 2),
        }
        for ln in report.lines
    ]
    return {"csv": write_summary_csv(rows, d / "excel_replenishment.csv")}


def build_deck(
    report: ExcelReplenishmentReport,
    *,
    client: str = "Client",
    prepared: str = "",
    citations: tuple[str, ...] = (),
    confidence: float = 0.85,
) -> Deliverable:
    """Compose the study: what is short, by how much, and how to act on the planilla."""
    short = [ln for ln in report.lines if ln.restock_qty > 0]
    findings = [
        Finding(
            f"{report.n_restock} SKU(s) below target",
            f"{report.total_restock:,.0f} units short across {report.n_skus} SKU(s) read from "
            f"{report.filename} ({report.sheet}, {report.mode} mode).",
            impact="replenish to avoid stockouts on the thin SKUs",
        )
    ]
    if short:
        worst = max(short, key=lambda ln: ln.restock_qty)
        findings.append(
            Finding(
                f"Thinnest SKU: {worst.sku}",
                f"{worst.on_hand:.0f} on hand vs a {worst.target:.0f} target (+{worst.restock_qty:.0f} needed).",
                impact="prioritize this replenishment line",
            )
        )
    kpis = (
        Kpi("SKUs read", str(report.n_skus), rationale=f"From {report.filename}"),
        Kpi("SKUs to replenish", str(report.n_restock), target="0", rationale="Below target"),
        Kpi("Units short", f"{report.total_restock:,.0f}", target="0", rationale="Restock to reach target"),
        Kpi(
            "Target rule",
            f"{report.cover_periods:.0f} periods of demand" if report.mode == "demand-cover"
            else f"{report.order_up_to_factor:.1f} x reorder point",
            rationale="How the target was set",
        ),
    )
    data_sources = (
        DataSource("Stock / reorder data", f"Client planilla {report.filename} ({report.sheet})", "on run"),
        DataSource("Order-quantity write-back", "Staged via src/connectors/excel.py (safe-staging plane)",
                   "on apply"),
    )
    recommendations = [
        "Approve and apply the staged order column so the planilla itself carries the plan.",
        "Keep the reorder/demand columns current - the plan is only as good as the planilla.",
    ]
    return Deliverable(
        title="Excel Replenishment",
        client=client,
        summary=report.summary,
        findings=tuple(findings),
        kpis=kpis,
        data_sources=data_sources,
        recommendations=tuple(recommendations),
        citations=tuple(citations),
        confidence=confidence,
        residual="Applying the staged order quantities writes to the client's file through the "
                 "Excel connector's safe-staging plane (drift check, backup, atomic write, "
                 "rollback); a human approves before anything is committed.",
        prepared=prepared,
    )
