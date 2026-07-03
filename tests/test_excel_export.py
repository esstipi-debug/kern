"""Tests for Excel workbook export."""

from pathlib import Path

from openpyxl import load_workbook

from src.excel_export import write_analysis_workbook


def test_write_analysis_workbook_sheets(tmp_path: Path):
    path = tmp_path / "test.xlsx"
    write_analysis_workbook(
        path,
        product_id="SKU-A",
        parameters={"D": 1000, "h": 1.75},
        results={"Q*": 239},
        gsm={"nodes": [], "total_holding_cost": 485, "echelon_order_up_to": (1, 2, 3)},
        simulation={"fill_rate": 0.95},
        newsvendor={"Q*": 4},
    )
    wb = load_workbook(path)
    assert "Summary" in wb.sheetnames
    assert "Parameters" in wb.sheetnames
    assert "Formulas" in wb.sheetnames
    assert "GSM" in wb.sheetnames
    assert wb["Summary"]["B5"].value == 239


def test_write_analysis_workbook_defuses_formula_injection(tmp_path: Path):
    """A result value starting with '=' must not be written as a live formula
    (OWASP CSV/Excel injection) - it must land as inert text when opened."""
    payload = '=cmd|" /C calc"!A0'
    path = tmp_path / "inj.xlsx"
    write_analysis_workbook(
        path,
        product_id="SKU-A",
        parameters={"D": 1000, "h": 1.75},
        results={"Q*": payload},
    )
    wb = load_workbook(path)
    cell = wb["Summary"]["B5"]
    assert cell.value != payload
    assert cell.data_type != "f"


def test_write_analysis_workbook_defuses_formula_injection_in_gsm_summary(tmp_path: Path):
    """The GSM sheet's total-holding-cost/echelon-levels rows are written via
    direct ws_gsm.cell() calls, bypassing _write_table() - same guard needed."""
    payload = '=cmd|" /C calc"!A0'
    path = tmp_path / "inj_gsm.xlsx"
    write_analysis_workbook(
        path,
        product_id="SKU-A",
        parameters={},
        results={},
        gsm={"nodes": [], "total_holding_cost": payload, "echelon_order_up_to": (1, 2, 3)},
    )
    wb = load_workbook(path)
    cell = wb["GSM"].cell(row=5, column=2)
    assert cell.value != payload
    assert cell.data_type != "f"
